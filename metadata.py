import os
import re
import sys
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from mutagen.easyid3 import EasyID3
from mutagen.id3 import ID3NoHeaderError


def build_result(file_path, status, reason="", artist="", title="", separator=""):
    """Build a normalized row for the final report."""
    return {
        "file_name": os.path.basename(file_path),
        "file_path": file_path,
        "status": status,
        "reason": reason,
        "artist": artist,
        "title": title,
        "separator": separator,
    }


def export_report_to_excel(report_rows, output_path):
    """Export processing results to a formatted Excel file."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Metadata Report"

    headers = [
        "File Name",
        "Full Path",
        "Status",
        "Reason",
        "Artist",
        "Title",
        "Separator Used",
    ]
    sheet.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    success_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
    failure_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")

    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in report_rows:
        sheet.append(
            [
                row["file_name"],
                row["file_path"],
                row["status"],
                row["reason"],
                row["artist"],
                row["title"],
                row["separator"],
            ]
        )

    for row_index in range(2, sheet.max_row + 1):
        status_cell = sheet.cell(row=row_index, column=3)
        status_text = str(status_cell.value).strip().lower()
        if status_text == "success":
            status_cell.fill = success_fill
        else:
            status_cell.fill = failure_fill

    for column_cells in sheet.columns:
        max_len = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        sheet.column_dimensions[column_letter].width = min(max_len + 2, 70)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:G{sheet.max_row}"

    workbook.save(output_path)


def insert_metadata_from_filename(file_path):
    """
        Extracts artist name and title from filename and writes them to MP3 metadata.

        Supported separators (in priority order):
        1. " - " (recommended)
        2. ":"
        3. "_"

        Fallback behavior:
        - If no supported separator is found, title becomes the full filename and
            artist defaults to "Unknown Artist".
    """
    try:
        # Ensure file exists and is an MP3.
        if not os.path.isfile(file_path) or not file_path.lower().endswith(".mp3"):
            print(f"Skipping: {file_path} (not an MP3 file)")
            return build_result(file_path, "Failure", "Not an MP3 file")

        # Extract filename without extension
        filename = os.path.splitext(os.path.basename(file_path))[0]
        separator_used = ""

        # Split into artist and title
        if " - " not in filename and ":" not in filename and "_" not in filename:
            print(f"Skipping separator parse: {filename} (no supported separator found; using fallback)")
            title = filename.strip()
            artist = "Unknown Artist"
            separator_used = "fallback"
        elif " - " in filename:
            artist, title = filename.split(" - ", 1)
            separator_used = " - "
        elif ":" in filename:
            artist, title = filename.split(":", 1)
            separator_used = ":"
        else:
            artist, title = filename.split("_", 1)
            separator_used = "_"

        # Remove known suffix markers without stripping all bracketed text.
        clean_title = title.strip()
        clean_title = re.sub(
            r"\s*[\[(]\s*official\s+(?:music\s+video|video|hd\s+video|song)\s*[\])]\s*",
            " ",
            clean_title,
            flags=re.IGNORECASE,
        )
        clean_title = re.sub(r"\s*\bofficial\s+video\b\s*", " ", clean_title, flags=re.IGNORECASE)
        clean_title = re.sub(r"\s*\[[A-Za-z0-9_-]{6,}\]\s*$", "", clean_title)
        clean_title = re.sub(r"\s{2,}", " ", clean_title).strip()

        # validate artist and title
        if not artist.strip() or not clean_title:
            print(f"Skipping: {filename} (artist or title is empty)")
            return build_result(
                file_path,
                "Failure",
                "Artist or title is empty after parsing/cleanup",
                artist.strip(),
                clean_title,
                separator_used,
            )


        # Load or create ID3 tags
        try:
            audio = EasyID3(file_path)
        except ID3NoHeaderError:
            audio = EasyID3()
            audio.save(file_path)  # Create empty tag
            audio = EasyID3(file_path)

        # Set metadata
        audio["artist"] = artist.strip()
        audio["title"] = clean_title
        try:
            audio.save()
        except PermissionError:
            print(f"Permission denied: {file_path} (close apps using this file or remove read-only)")
            return build_result(
                file_path,
                "Failure",
                "Permission denied while saving tags",
                artist.strip(),
                clean_title,
                separator_used,
            )

        print(f"Updated: {file_path} → Artist '{artist.strip()}', Title '{clean_title}'")
        return build_result(
            file_path,
            "Success",
            "",
            artist.strip(),
            clean_title,
            separator_used,
        )

    except PermissionError:
        print(f"Permission denied: {file_path} (close apps using this file or remove read-only)")
        return build_result(file_path, "Failure", "Permission denied")

    except Exception as e:
        print(f"Error processing {file_path}: {e}")
        return build_result(file_path, "Failure", f"Unexpected error: {e}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python insert_metadata.py <mp3_folder_path>")
        sys.exit(1)

    folder_path = sys.argv[1]

    if not os.path.isdir(folder_path):
        print("Invalid folder path.")
        sys.exit(1)

    # Process files and export processing report.
    report_rows = []
    for file_name in os.listdir(folder_path):
        file_path = os.path.join(folder_path, file_name)
        report_rows.append(insert_metadata_from_filename(file_path))

    report_path = os.path.join(folder_path, "metadata_report.xlsx")
    export_report_to_excel(report_rows, report_path)

    success_count = sum(1 for row in report_rows if row["status"] == "Success")
    failure_count = len(report_rows) - success_count
    print(f"\nReport exported: {report_path}")
    print(f"Summary: {success_count} success, {failure_count} failure")
